import argparse
import openpyxl
import json
import os

extra_field_names = ["description", "tag", "conformityTopic", "status", "thresholdValue", "performanceLevel", "category"]

def parse_excel_to_svc_json(
    workbook,
    sheet_name: str,
    max_level: int,
    levels: int = 1
):
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found.")

    ws = workbook[sheet_name]
    current_levels = [None] * levels
    tree = []

    for row in ws.iter_rows(max_col=max_level + 1 + len(extra_field_names)):
        extra_data = {}
        for i, field_name in enumerate(extra_field_names):
            col_idx = max_level + 1 + i
            value = row[col_idx].value if col_idx < len(row) else None

            if field_name == 'tag' and value:
                value = [item.strip() for item in value.split(",\n")]

            if value:
                extra_data[field_name] = value

        for level in range(levels):
            id_cell = row[level].value
            name_cell = row[level + 1].value

            if id_cell:
                if str(id_cell).strip() == 'ID':
                    continue
                node = {
                    "type": ["Criterion"],
                    "id": str(id_cell).strip(),
                    "name": str(name_cell).strip() if name_cell else ""
                }
                node.update(extra_data)
                node["subCriterion"] = []

                current_levels[level] = node
                for j in range(level + 1, levels):
                    current_levels[j] = None

                if level == 0:
                    tree.append(node)
                else:
                    parent = current_levels[level - 1]
                    if parent:
                        parent["subCriterion"].append(node)

                break

    return tree

def detect_max_levels_by_column_count(wb, sheet_name):
    """Calculate levels based on known extra fields and total columns."""
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found.")

    ws = wb[sheet_name]
    max_columns = ws.max_column
    return (max_columns - len(extra_field_names) - 1)

def main():
    parser = argparse.ArgumentParser(description="Excel to Nested JSON Converter")
    parser.add_argument('--input', help='Path to Excel file (.xlsx)')
    parser.add_argument('--sheet', help='Sheet name to parse')
    parser.add_argument('--levels', type=int, help='Number of levels to map (overrides auto-detection)')
    parser.add_argument('--output', help='Path to output JSON file')

    args = parser.parse_args()

    print("=== Excel to Nested JSON Converter ===")

    # Input path
    input_path = args.input or input("Enter input Excel file path (.xlsx): ").strip()
    while not os.path.isfile(input_path):
        input_path = input("File not found. Please enter a valid file path (.xlsx): ").strip()

    try:
        wb = openpyxl.load_workbook(input_path)
    except Exception as e:
        print(f"❌ Failed to open Excel file: {e}")
        return

    # Sheet name
    available_sheets = wb.sheetnames
    sheet_name = args.sheet or input(f"Enter sheet name to parse (available: {', '.join(available_sheets)}): ").strip()
    while sheet_name not in available_sheets:
        print(f"❌ Sheet '{sheet_name}' not found.")
        sheet_name = input(f"Please enter a valid sheet name (available: {', '.join(available_sheets)}): ").strip()

    # Detect max level
    max_level = detect_max_levels_by_column_count(wb, sheet_name)
    print(f"Detected {max_level} levels based on column structure.")

    # Levels to use
    if args.levels is not None:
        levels = args.levels
    else:
        level_input = input(f"Enter number of levels to map [default {max_level}]: ").strip()
        if level_input:
            try:
                levels = int(level_input)
            except ValueError:
                print("Invalid number. Defaulting to detected level.")
                levels = max_level
        else:
            levels = max_level

    # Output path
    output_path = args.output or input("Enter output JSON file path: ").strip()
    if not output_path.lower().endswith(".json"):
        output_path += ".json"

    try:
        json_result = parse_excel_to_svc_json(
            workbook=wb,
            sheet_name=sheet_name,
            max_level=max_level,
            levels=levels
        )

        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(json_result, f, indent=2, ensure_ascii=False)

        print(f"\n✅ JSON exported to: {output_path}")
    except Exception as e:
        print(f"\n❌ Error: {e}")



if __name__ == "__main__":
    main()
